"""
文档解析器 - 支持 PDF、Word、Excel、图片
"""

import os
import tempfile
from typing import List


class DocumentParser:
    """文档解析器，将各种文档转为图片列表"""
    
    @staticmethod
    def parse_to_images(file_path: str) -> List[str]:
        """
        将文档转换为图片列表
        返回图片路径列表
        """
        ext = os.path.splitext(file_path)[1].lower()
        
        # 图片直接返回
        if ext in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff']:
            return [file_path]
        
        # PDF 转图片
        elif ext == '.pdf':
            return DocumentParser._pdf_to_images(file_path)
        
        # Word 转图片（提取所有内嵌图片）
        elif ext in ['.doc', '.docx']:
            return DocumentParser._word_to_images(file_path)
        
        # Excel 转图片
        elif ext in ['.xls', '.xlsx']:
            return DocumentParser._excel_to_images(file_path)
        
        else:
            raise ValueError(f"不支持的文件类型: {ext}")
    
    @staticmethod
    def _pdf_to_images(pdf_path: str) -> List[str]:
        """PDF转图片 - 使用 PyMuPDF（不需要 poppler）"""
        try:
            import fitz  # PyMuPDF
            from PIL import Image
            import io
            
            doc = fitz.open(pdf_path)
            images = []
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                
                # 方法1：提取页面中的图片
                image_list = page.get_images()
                for img_index, img in enumerate(image_list):
                    xref = img[0]
                    base_image = doc.extract_image(xref)
                    image_data = base_image["image"]
                    img_pil = Image.open(io.BytesIO(image_data))
                    
                    temp_path = tempfile.NamedTemporaryFile(suffix=f'_pdf_page_{page_num+1}_img_{img_index}.png', delete=False).name
                    img_pil.save(temp_path, 'PNG')
                    images.append(temp_path)
                    print(f"PDF第{page_num+1}页提取图片")
                
                # 方法2：如果没有图片，把整页转成图片
                if not image_list:
                    pix = page.get_pixmap(dpi=150)
                    img_data = pix.tobytes("png")
                    img_pil = Image.open(io.BytesIO(img_data))
                    
                    temp_path = tempfile.NamedTemporaryFile(suffix=f'_pdf_page_{page_num+1}.png', delete=False).name
                    img_pil.save(temp_path, 'PNG')
                    images.append(temp_path)
                    print(f"PDF第{page_num+1}页转换为图片")
            
            doc.close()
            return images
        except Exception as e:
            print(f"PDF转换失败: {e}")
            return []
    
    @staticmethod
    def _word_to_images(doc_path: str) -> List[str]:
        """Word转图片 - 提取所有内嵌图片"""
        try:
            from docx import Document
            from PIL import Image
            import io
            
            doc = Document(doc_path)
            images = []
            image_count = 0
            
            # 提取 Word 中的所有内嵌图片
            for rel in doc.part.rels.values():
                if "image" in rel.target_ref:
                    try:
                        image_data = rel.target_part.blob
                        img = Image.open(io.BytesIO(image_data))
                        
                        temp_path = tempfile.NamedTemporaryFile(suffix=f'_word_img_{image_count}.png', delete=False).name
                        img.save(temp_path, 'PNG')
                        images.append(temp_path)
                        image_count += 1
                        print(f"提取到 Word 内嵌图片 {image_count}")
                    except Exception as e:
                        print(f"提取图片失败: {e}")
            
            # 如果没有内嵌图片，提取文本转图片
            if image_count == 0:
                text = '\n'.join([para.text for para in doc.paragraphs if para.text.strip()])
                if text:
                    from PIL import Image, ImageDraw, ImageFont
                    
                    img = Image.new('RGB', (1000, 800), color='white')
                    draw = ImageDraw.Draw(img)
                    
                    try:
                        font = ImageFont.truetype("simsun.ttc", 16)
                    except:
                        font = ImageFont.load_default()
                    
                    y = 10
                    for line in text.split('\n')[:50]:
                        draw.text((10, y), line[:120], fill='black', font=font)
                        y += 25
                    
                    temp_path = tempfile.NamedTemporaryFile(suffix='_word_text.png', delete=False).name
                    img.save(temp_path, 'PNG')
                    images.append(temp_path)
                    print(f"Word文本转换为图片")
            
            if not images:
                print(f"Word文件无内容: {doc_path}")
            
            return images
            
        except Exception as e:
            print(f"Word转换失败: {e}")
            return []
    
    @staticmethod
    def _excel_to_images(excel_path: str) -> List[str]:
        """Excel转图片"""
        try:
            import openpyxl
            from PIL import Image, ImageDraw, ImageFont
            
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            images = []
            
            # 处理所有工作表
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                sheet = wb[sheet_name]
                
                text_lines = []
                for row in sheet.iter_rows(max_row=50, values_only=True):
                    row_text = '\t'.join([str(cell) if cell else '' for cell in row])
                    if row_text.strip():
                        text_lines.append(row_text)
                
                text = '\n'.join(text_lines[:100])
                
                if text:
                    img = Image.new('RGB', (1200, 900), color='white')
                    draw = ImageDraw.Draw(img)
                    
                    try:
                        font = ImageFont.truetype("simsun.ttc", 14)
                    except:
                        font = ImageFont.load_default()
                    
                    y = 10
                    for line in text.split('\n')[:60]:
                        draw.text((10, y), line[:150], fill='black', font=font)
                        y += 20
                    
                    temp_path = tempfile.NamedTemporaryFile(suffix=f'_excel_sheet_{sheet_idx}.png', delete=False).name
                    img.save(temp_path, 'PNG')
                    images.append(temp_path)
                    print(f"Excel工作表 {sheet_name} 转换为图片")
            
            return images
        except Exception as e:
            print(f"Excel转换失败: {e}")
            return []


# 测试
if __name__ == "__main__":
    result = DocumentParser.parse_to_images("test.docx")
    print(f"解析到 {len(result)} 张图片")
    for r in result:
        print(f"  - {r}")